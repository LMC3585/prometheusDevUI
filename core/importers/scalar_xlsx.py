"""Importer for the Scalar course design spreadsheet template.

The template is column oriented with the following required headers:
Course Name, CLO (or CLOs), Topic, Subtopic, Lesson Title, Performance
Criteria. Each row represents the most granular relationship between a
lesson and a performance criteria aligned to a CLO. This importer loads
that data into the canonical course models.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from core.models.course import (
    CLO,
    CourseMetadata,
    Lesson,
    PerformanceCriteria,
    Subtopic,
    Topic,
)

ScalarRow = Dict[str, Optional[str]]

HEADER_ALIASES = {
    "course name": "course_name",
    "course title": "course_name",
    "clo": "clo",
    "clos": "clo",
    "learning outcome": "clo",
    "topic": "topic",
    "topics": "topic",
    "subtopic": "subtopic",
    "sub-topics": "subtopic",
    "lesson": "lesson_title",
    "lesson title": "lesson_title",
    "lessons": "lesson_title",
    "performance criteria": "performance_criteria",
    "pc": "performance_criteria",
    "pcs": "performance_criteria",
}

REQUIRED_FIELDS = {
    "course_name",
    "clo",
    "topic",
    "subtopic",
    "lesson_title",
    "performance_criteria",
}


def load_course_from_scalar(path: str | Path) -> CourseMetadata:
    """Load a Scalar XLSX file into the canonical `CourseMetadata` graph."""

    parser = _ScalarWorkbookParser(Path(path))
    return parser.parse()


@dataclass
class _ScalarWorkbookParser:
    workbook_path: Path

    def parse(self) -> CourseMetadata:
        wb = load_workbook(self.workbook_path, data_only=True)
        ws = wb.active
        if ws.max_row < 2:
            raise ValueError("Scalar workbook does not contain any data rows.")

        header_map = self._resolve_headers(ws[1])
        normalized_rows = list(
            filter(
                None,
                (self._normalize_row(row, header_map) for row in ws.iter_rows(min_row=2, values_only=True)),
            )
        )
        if not normalized_rows:
            raise ValueError("Scalar workbook rows are empty after normalization.")

        builder = _CourseBuilder()
        for row in normalized_rows:
            builder.ingest(row)

        return builder.build()

    def _resolve_headers(self, header_row) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            raw_header = str(cell.value or "").strip().lower()
            if not raw_header:
                continue
            normalized = HEADER_ALIASES.get(raw_header)
            if normalized:
                header_map[normalized] = idx

        missing = [field for field in REQUIRED_FIELDS if field not in header_map]
        if missing:
            raise ValueError(f"Scalar workbook missing required headers: {', '.join(missing)}")
        return header_map

    def _normalize_row(self, row_values, header_map: Dict[str, int]) -> Optional[ScalarRow]:
        normalized: ScalarRow = {}
        empty = True
        for field_name, column_idx in header_map.items():
            value = row_values[column_idx] if column_idx < len(row_values) else None
            if isinstance(value, str):
                value = value.strip()
            if value not in (None, ""):
                empty = False
            normalized[field_name] = value or None

        return None if empty else normalized


class _CourseBuilder:
    """Incrementally constructs the in-memory course graph."""

    def __init__(self) -> None:
        self.course_title: Optional[str] = None
        self.clos: Dict[str, CLO] = {}
        self.topics: Dict[str, Topic] = {}
        self.subtopics: Dict[Tuple[str, str], Subtopic] = {}
        self.lessons: Dict[Tuple[str, str, str], Lesson] = {}
        self.performance_criteria: Dict[str, PerformanceCriteria] = {}
        self.topic_order = 1
        self.subtopic_order: Dict[str, int] = {}

    def ingest(self, row: ScalarRow) -> None:
        course_title = row.get("course_name")
        if course_title and not self.course_title:
            self.course_title = course_title

        clo_text = row.get("clo")
        topic_text = row.get("topic")
        subtopic_text = row.get("subtopic")
        lesson_title = row.get("lesson_title")
        pc_text = row.get("performance_criteria")

        if not all([clo_text, topic_text, subtopic_text, lesson_title]):
            raise ValueError(
                "Each Scalar row must include CLO, Topic, Subtopic, and Lesson Title. "
                f"Got: {row}"
            )

        clo = self._get_or_create_clo(clo_text)
        topic = self._get_or_create_topic(topic_text)
        self._ensure_list_membership(topic.associated_clos, clo.key)

        subtopic = self._get_or_create_subtopic(topic, subtopic_text)
        self._ensure_list_membership(subtopic.associated_clos, clo.key)

        lesson = self._get_or_create_lesson(subtopic, lesson_title)
        self._ensure_list_membership(lesson.associated_clos, clo.key)

        if pc_text:
            pc = self._get_or_create_performance_criteria(pc_text)
            self._attach_pc(lesson, clo, pc)

    def build(self) -> CourseMetadata:
        title = self.course_title or "Untitled Course"
        course_id = _slugify(f"course-{title}")

        return CourseMetadata(
            course_id=course_id,
            title=title,
            clos=list(self.clos.values()),
            topics=list(self.topics.values()),
        )

    def _get_or_create_clo(self, text: str) -> CLO:
        key = _slugify(f"clo-{text}")
        if key not in self.clos:
            self.clos[key] = CLO(key=key, description=text)
        return self.clos[key]

    def _get_or_create_topic(self, text: str) -> Topic:
        key = _slugify(f"topic-{text}")
        if key not in self.topics:
            topic = Topic(key=key, title=text, order=self.topic_order)
            self.topic_order += 1
            self.topics[key] = topic
        return self.topics[key]

    def _get_or_create_subtopic(self, topic: Topic, text: str) -> Subtopic:
        key = _slugify(f"subtopic-{text}")
        compound_key = (topic.key, key)
        if compound_key not in self.subtopics:
            order = self.subtopic_order.get(topic.key, 1)
            subtopic = Subtopic(key=key, title=text, order=order)
            self.subtopics[compound_key] = subtopic
            self.subtopic_order[topic.key] = order + 1
            topic.subtopics.append(subtopic)
        return self.subtopics[compound_key]

    def _get_or_create_lesson(self, subtopic: Subtopic, title: str) -> Lesson:
        key = _slugify(f"lesson-{title}")
        compound_key = (subtopic.key, key, title)
        if compound_key not in self.lessons:
            lesson = Lesson(key=key, title=title)
            self.lessons[compound_key] = lesson
            subtopic.lessons.append(lesson)
        return self.lessons[compound_key]

    def _get_or_create_performance_criteria(self, text: str) -> PerformanceCriteria:
        key = _slugify(f"pc-{text}")
        if key not in self.performance_criteria:
            self.performance_criteria[key] = PerformanceCriteria(key=key, description=text)
        return self.performance_criteria[key]

    @staticmethod
    def _attach_pc(lesson: Lesson, clo: CLO, pc: PerformanceCriteria) -> None:
        if not any(existing.key == pc.key for existing in lesson.performance_criteria):
            lesson.performance_criteria.append(pc)
        if not any(existing.key == pc.key for existing in clo.performance_criteria):
            clo.performance_criteria.append(pc)

    @staticmethod
    def _ensure_list_membership(container: List[str], value: str) -> None:
        if value not in container:
            container.append(value)


_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


def _slugify(value: str) -> str:
    value = value.strip().lower()
    value = _NON_ALNUM_RE.sub("-", value).strip("-")
    return value or "item"

